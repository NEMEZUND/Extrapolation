# Импортируем необходимые библиотеки
import openpyxl
import PySimpleGUI as sg

# Определяем функцию для заполнения пустых ячеек в столбце
def fill_empty_cells(sheet, column_index, max_empty_cells=20):
    # Получаем общее количество строк в листе
    total_rows = sheet.max_row
    row = 1

    # Итерируемся по каждой строке
    while row <= total_rows:
        # Получаем значение ячейки в указанном столбце
        cell_value = sheet.cell(row=row, column=column_index).value

        # Проверяем, содержит ли ячейка значение
        if cell_value:
            current_row = row + 1
            empty_cells_count = 0

            # Итерируемся по пустым ячейкам под текущим населенным пунктом
            while empty_cells_count < max_empty_cells and current_row <= total_rows:
                if not sheet.cell(row=current_row, column=column_index).value:
                    sheet.cell(row=current_row, column=column_index).value = cell_value
                    empty_cells_count += 1
                else:
                    break

                current_row += 1

            row = current_row
        else:
            row += 1

        # Обновляем процентный индикатор выполнения
        current_percentage = int((row / total_rows) * 100)
        window['-PROGRESS-'].update_bar(current_percentage)

# Определяем макет интерфейса
layout = [
    [sg.Text('Выберите файл Excel')],
    [sg.Input(key='-INPUT-', enable_events=True), sg.FileBrowse()],
    [sg.Text('Укажите лист')],
    [sg.InputText(key='-SHEET-')],
    [sg.Text('Укажите номер столбца')],
    [sg.InputText(key='-COLUMN-', size=(5, 1))],
    [sg.Text('Выберите путь для сохранения')],
    [sg.Input(key='-OUTPUT-', enable_events=True), sg.FolderBrowse()],
    [sg.ProgressBar(100, orientation='h', size=(20, 20), key='-PROGRESS-')],
    [sg.Button('Выполнить')],
]

# Создаем окно интерфейса
window = sg.Window('Fill Empty Cells', layout, finalize=True)

# Запускаем бесконечный цикл для обработки событий
while True:
    event, values = window.read()

    # Проверяем событие закрытия окна
    if event == sg.WINDOW_CLOSED:
        break
    # Проверяем событие выполнения операции
    elif event == 'Выполнить':
        # Получаем пути и значения из полей ввода
        file_path = values['-INPUT-']
        sheet_name = values['-SHEET-']

        try:
            # Пытаемся преобразовать введенный номер столбца в целое число
            column_index = int(values['-COLUMN-'])
        except ValueError:
            sg.popup_error('Введите корректный номер столбца (целое число).')
            continue

        try:
            # Загружаем рабочую книгу и выбранный лист
            wb = openpyxl.load_workbook(file_path)
            sheet = wb[sheet_name]
        except Exception as e:
            # Выводим сообщение об ошибке, если что-то идет не так
            sg.popup_error(f'Ошибка при открытии файла: {e}')
            continue

        # Вызываем функцию для заполнения пустых ячеек
        fill_empty_cells(sheet, column_index)

        try:
            # Получаем путь для сохранения файла и имя файла от пользователя
            output_path = values['-OUTPUT-']
            output_file_name = sg.popup_get_text('Введите имя файла для сохранения (без расширения)')
            if output_file_name:
                output_file_path = f"{output_path}/{output_file_name}.xlsx"
                # Сохраняем изменения и выводим сообщение об успехе
                wb.save(output_file_path)
                sg.popup('Операция выполнена успешно!', title='Успех')
        except Exception as e:
            sg.popup_error(f'Ошибка при сохранении файла: {e}')

# Закрываем окно после завершения цикла
window.close()
